"""
Дан CVS (ready.csv) строк бывает разное количество. Так же дан XLSX (master.xlsx).
Нужно данный CVS перенести на первый лист XLSX не потеряв структуры.

Для работы установлена библиотека openpyxl
"""
from openpyxl import load_workbook
import PySimpleGUI as sg

sg.theme('DarkGreen')
print = sg.Print

print('*** Start ***')

layout = [[sg.Text('Выбери файлы'), sg.Image(sg.EMOJI_BASE64_HAPPY_THUMBS_UP)],
          [sg.Text('XLSX'), sg.InputText(key = 'xlsx_f'), sg.FileBrowse(file_types = (('XLSX Files', '.XLSX'),))],
          [sg.Text('CSV'), sg.InputText(key = 'cvs_f'), sg.FileBrowse(file_types = (('CSV Files', '.csv'),))],
          [sg.B('Ebosh'), sg.Cancel()]]

window = sg.Window('Huyarim CVS -> XLSX', layout)

event, values = window.read()
print(f'XLSX: {values["xlsx_f"]} \nCVS: {values["cvs_f"]}')

window.close()


save_f = sg.popup_get_file('Save finished file as', save_as = True, file_types = (('XLSX Files', '.XLSX'),))

print(f'Save as = {save_f}')



def do_cvs(cvs_l:list) -> list:
    """Функция принимает список и преобразоует часть элементов в нужный формат.
    Возвращает список"""
    if not cvs_l[0]: return ['']
    else:
        int_col = (0, 8, 22, 25, 28, 45) # номера столбцов с целыми числами (нумерация с 0)
        for i in int_col:
            if cvs_l[i]: cvs_l[i] = int(cvs_l[i])

        float_col = (23,) # номера столбцов с десятичными числами (нумерация с 0)
        for i in float_col:
            cvs_l[i] = float(cvs_l[i].replace(',', '.'))

        strip_col = (7,) # номера столбцов, где неадо удалить кавычки (нумерация с 0)
        for i in strip_col:
            cvs_l[i] = str(cvs_l[i]).strip('"')

        strip_col2 = (16, 19) # номера столбцов где нао удалить пробелы сначала и конца строки (нумерация с 0)
        for i in strip_col2:
            cvs_l[i] = cvs_l[i].strip()

        return cvs_l
def do_data(columns: tuple, start_index: int, stop_index: int) -> None:
    """Функция примет ортеж с перечислением нужных столбцов и поменяет формат
    их данных на строковый формата DD.MM.YYYY"""
    for cn in columns:
        for i in range(start_index, stop_index):
            x = str(ws[cn + str(i)].value)
            ws[cn + str(i)].value = f'{x[8:10]}.{x[5:7]}.{x[0:4]}'


# работаем с .CVS
f_cvs = open(values['cvs_f'], encoding= 'utf8')
my_cvs = do_cvs(f_cvs.readline().split(sep = ';'))

# работаем с .XLSX
wb = load_workbook(values["xlsx_f"])
ws = wb[wb.sheetnames[0]]
print(f'Работаем с листом {wb.sheetnames[0]}')

# Определяем где начинается пустота
for i in range(1, ws.max_row):
    cell = ws.cell(row= i, column=1)
    if not cell.value:
        print(f"На {i} строке начнем вставлять CVS.")
        break
ws.delete_rows(i, ws.max_row-i+1)

# форматируем ячейки с датой
data_col = ('E', 'AQ', 'AR') # имена столбцов (как в итоговой таблице) в которых должна быть ДАТА
do_data(data_col, 2, i)

# Вставляем строки CVS
i = 0
while my_cvs[0]:
    ws.append(my_cvs)
    my_cvs = do_cvs(f_cvs.readline().split(sep=';'))
    i += 1
print(f'Вставили {i} строк.')

wb.save(save_f)
print(f'Сохранили всё в {save_f}')
print('***  Всё ОК! ***\n')
